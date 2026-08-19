"""Client condiviso per le API pubbliche "PCF" (estrazione curve/letture).

Protocollo comune a Duereti e Unareti (stesso schema requestToken ->
requestExport -> requestResult, verificato identico tra i due manuali),
parametrizzato sul solo base_url. Le osservazioni specifiche sul
comportamento del WAF riportate nei commenti sotto sono state fatte su
Duereti; se emergono differenze su Unareti vanno annotate qui.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import logging
import re
import time
import zipfile
from dataclasses import dataclass, field
from datetime import date, timedelta

import aiohttp

from .const import (
    ESITO_OK,
    MAX_DATE_RANGE_MONTHS,
    MAX_SUPPLY_POINTS_PER_REQUEST,
    MODE_CURVE,
    RESULT_POLL_INTERVAL_SECONDS,
    RITARDO_VERIFICA_POD_GIORNI,
    RESULT_POLL_MAX_ATTEMPTS,
    TOKEN_SAFETY_MARGIN_SECONDS,
    TOKEN_VALIDITY_SECONDS,
)

_LOGGER = logging.getLogger(__name__)

# Chiavi il cui valore non va MAI loggato in chiaro: le credenziali API e il
# token di sessione. home-assistant.log viene spesso condiviso per chiedere
# supporto, quindi qualunque cosa finisca qui va considerata pubblica.
_CHIAVI_SEGRETE = {"clientid", "secretid", "authorization", "token"}


def _oscura(valore: object) -> str:
    """Riduce un segreto a una forma che resta utile per il debug (si capisce
    se è presente, se cambia tra una chiamata e l'altra, quanto è lungo) senza
    rivelarlo."""
    testo = str(valore)
    if not testo:
        return "<vuoto>"
    if len(testo) <= 8:
        return f"<{len(testo)} caratteri>"
    return f"{testo[:4]}…{testo[-2:]} (<{len(testo)} caratteri>)"


def _sanitizza(dati: dict | None) -> dict:
    """Copia di un dict con i valori sensibili oscurati, pronta per il log."""
    if not dati:
        return {}
    return {
        chiave: (_oscura(valore) if chiave.lower() in _CHIAVI_SEGRETE else valore)
        for chiave, valore in dati.items()
    }


def descrivi_errore(err: Exception, distributor_display_name: str = "del distributore") -> str:
    """Rende leggibile un errore delle API, in particolare il blocco del WAF.

    Quando il WAF (F5 BIG-IP) rifiuta una richiesta restituisce una pagina
    HTML al posto del JSON: mostrarla intera all'utente è inutile e
    incomprensibile. Ne estraiamo il solo Support ID, che è l'unica
    informazione utile (serve al distributore per rintracciare il blocco nei log).

    Il comportamento del WAF descritto qui è stato osservato e verificato su
    Duereti. Non ancora riconfermato su Unareti (stessa infrastruttura
    dichiarata, ma da verificare sul campo se emergono blocchi analoghi):
    'distributor_display_name' serve a non attribuire per errore un
    comportamento non ancora verificato al distributore sbagliato nel
    messaggio mostrato all'utente.
    """
    testo = str(err)
    if "Request Rejected" not in testo:
        return testo
    match = re.search(r"support ID is:\s*(\d+)", testo, re.IGNORECASE)
    support_id = match.group(1) if match else "non disponibile"
    return (
        f"richiesta rifiutata dal WAF {distributor_display_name} (blocco intermittente, "
        f"non dipende dalle credenziali). Support ID: {support_id}"
    )


def _log_richiesta(url: str, body: dict | None, headers: dict | None) -> None:
    """Logga la richiesta in uscita (solo a livello DEBUG).

    Serve a confrontare byte per byte cosa manda l'integrazione rispetto a un
    client che funziona (curl/Bruno), utile per capire perché il WAF di
    Duereti rifiuta le une e non le altre.
    """
    if not _LOGGER.isEnabledFor(logging.DEBUG):
        return
    _LOGGER.debug(
        "--> POST %s\n    headers: %s\n    body: %s",
        url,
        _sanitizza(headers),
        _sanitizza(body),
    )


def _log_risposta(url: str, status: int, headers: object, corpo: str) -> None:
    """Logga la risposta grezza (solo a livello DEBUG).

    Il corpo viene loggato così com'è (troncato): quando il WAF interviene non
    è JSON ma una pagina HTML "Request Rejected", e vedere il testo esatto con
    il Support ID è proprio l'informazione che serve.
    """
    if not _LOGGER.isEnabledFor(logging.DEBUG):
        return
    anteprima = corpo if len(corpo) <= 800 else f"{corpo[:800]}… [troncato, {len(corpo)} caratteri]"
    _LOGGER.debug(
        "<-- %s da %s\n    headers: %s\n    body: %s",
        status,
        url,
        dict(headers) if headers else {},
        anteprima,
    )


class PcfApiError(Exception):
    """Errore generico nella comunicazione con le API Duereti."""

    def __init__(self, message: str, http_status: int | None = None, data: dict | None = None) -> None:
        super().__init__(message)
        self.http_status = http_status
        self.data = data or {}


class PcfAuthError(PcfApiError):
    """401 UNAUTHORIZED - credenziali o token non validi/scaduti."""


class PcfValidationError(PcfApiError):
    """400 BAD_REQUEST - validazione input fallita (dettagli in message)."""


class PcfConflictError(PcfApiError):
    """409 CONFLICT - token già esistente per l'utente (requestToken)."""


class PcfRateLimitError(PcfApiError):
    """429 - limite di richieste attive raggiunto."""


class PcfNotFoundError(PcfApiError):
    """404 NOT_FOUND - ticket presente ma non collegato a nessun dato."""


@dataclass
class CurvaPunto:
    """Un singolo punto curva (timestamp + valore)."""

    timestamp: "datetime"  # noqa: F821 - import sotto per evitare cicli
    valore_kwh: float
    # Valore grezzo della colonna FL_ORA_LEGALE del CSV. Serve a disambiguare
    # l'ora ripetuta al passaggio da ora legale a ora solare: il timestamp da
    # solo è ambiguo, questo flag dice a quale delle due ore appartiene.
    # Vedi statistics._offset_da_flag per l'interpretazione.
    ora_legale: str | None = None


@dataclass
class RisultatoLetture:
    """Risultato del parsing dello zip per un singolo POD."""

    pod: str
    punti: list = field(default_factory=list)  # list[CurvaPunto]


class PcfApiClient:
    """Wrapper asincrono per il flusso requestToken -> requestExport -> requestResult."""

    def __init__(
        self,
        session: aiohttp.ClientSession,
        client_id: str,
        secret_id: str,
        base_url: str,
    ) -> None:
        """base_url e' l'URL base delle API del distributore (es.
        "https://areaclienti.duereti.it/ClientiDueRetiWeb/public/misure"),
        da cui si derivano i tre endpoint. E' parametrico perche' questo
        client e' condiviso tra distributori diversi (stesso protocollo
        PCF, host diverso): vedi distributors/duereti.py e
        distributors/unareti.py."""
        self._session = session
        self._client_id = client_id
        self._secret_id = secret_id
        self._base_url = base_url.rstrip("/")
        self._url_token = f"{self._base_url}/requestToken"
        self._url_export = f"{self._base_url}/requestExport"
        self._url_result = f"{self._base_url}/requestResult"
        self._token: str | None = None
        self._token_expiry: float = 0.0

    async def async_assicura_token(self) -> None:
        """Si assicura di avere un token valido, riusando quello in cache.

        A differenza di async_validate_credentials non forza il rinnovo: serve
        al coordinator per sapere se la prima chiamata (requestToken) funziona,
        senza sprecare richieste verso un'API che le blocca a intermittenza.
        """
        await self._get_token()

    async def async_validate_credentials(self) -> None:
        """Verifica che Client ID/Secret ID siano validi.

        Solleva PcfAuthError se le credenziali sono sbagliate, o
        PcfApiError per qualunque altro problema di connessione. Non
        restituisce nulla se la validazione ha successo. Metodo pubblico
        pensato per config_flow.py (evita di dover chiamare _get_token,
        un metodo interno, da fuori la classe).
        """
        await self._get_token(force_refresh=True)

    @staticmethod
    def _token_dal_messaggio(messaggio: str | None) -> str | None:
        """Estrae il token dal messaggio di conflitto di requestToken.

        Quando esiste già un token attivo per l'utente, Duereti risponde 409
        CONFLICT e non ne emette uno nuovo: restituisce quello esistente
        dentro il testo del messaggio, nella forma:

            "Esiste già un token registrato dall'utente, token: XMzKH..."

        Recuperarlo da lì risolve alla radice il caso del riavvio di Home
        Assistant entro i 10 minuti di validità di un token precedente: la
        nuova istanza non ha nulla in cache, ma può comunque riusare quello
        ancora valido invece di restare bloccata ad aspettarne la scadenza.
        """
        if not messaggio:
            return None
        trovato = re.search(r"token\s*[:=]\s*([A-Za-z0-9+/_=-]{8,})", messaggio)
        return trovato.group(1) if trovato else None

    def _memorizza_token(self, token: str, scadenza_intera: bool = True) -> str:
        """Salva il token in cache con la relativa scadenza.

        Per un token recuperato da un messaggio di conflitto non sappiamo da
        quanto sia già in uso: gli attribuiamo una validità residua prudente
        invece dei 10 minuti pieni, così lo rinnoviamo prima che scada.
        """
        self._token = token
        durata = TOKEN_VALIDITY_SECONDS if scadenza_intera else TOKEN_SAFETY_MARGIN_SECONDS * 4
        self._token_expiry = time.monotonic() + durata - TOKEN_SAFETY_MARGIN_SECONDS
        return token

    async def _get_token(self, force_refresh: bool = False) -> str:
        """Ottiene un token valido, richiedendone uno nuovo se scaduto o mancante.

        Se ne esiste già uno attivo lato server, Duereti risponde 409 CONFLICT
        (verificato) e mette il token attivo dentro il testo del messaggio,
        invece di emetterne uno nuovo: lo estraiamo e lo riusiamo.
        """
        now = time.monotonic()
        if not force_refresh and self._token and now < self._token_expiry:
            return self._token

        payload = {"clientId": self._client_id, "secretId": self._secret_id}

        try:
            _log_richiesta(self._url_token, payload, None)
            async with self._session.post(self._url_token, json=payload) as resp:
                data = await self._json_or_raise(resp)
        except PcfValidationError as err:
            # Su requestToken un 400 significa credenziali mancanti
            # (verificato: "ClientID mancante"). È un problema di
            # configurazione, non di validazione dei dati richiesti: lo
            # trattiamo come errore di autenticazione, così Home Assistant
            # propone di reinserire le credenziali invece di limitarsi a
            # segnalare un aggiornamento fallito.
            raise PcfAuthError(str(err), http_status=400, data=err.data) from err
        except PcfConflictError as err:
            # 409: il token esistente è nel messaggio dell'errore.
            token = self._token_dal_messaggio(str(err))
            if token:
                _LOGGER.debug("requestToken in conflitto: riuso il token indicato nel messaggio")
                return self._memorizza_token(token, scadenza_intera=False)
            if self._token:
                _LOGGER.debug("requestToken in conflitto: riuso il token già in cache")
                self._token_expiry = now + TOKEN_SAFETY_MARGIN_SECONDS
                return self._token
            raise

        if data.get("esito") != ESITO_OK:
            # Il conflitto arriva come 409 (gestito sopra), ma controlliamo il
            # messaggio anche qui per prudenza: se un giorno lo restituissero
            # con HTTP 200, senza questo verrebbe scambiato per un errore di
            # credenziali e farebbe scattare una riautenticazione inutile.
            token = self._token_dal_messaggio(data.get("message"))
            if token:
                _LOGGER.debug(
                    "requestToken: esiste già un token attivo, lo riuso da messaggio"
                )
                return self._memorizza_token(token, scadenza_intera=False)
            raise PcfAuthError(data.get("message") or "Autenticazione fallita")

        token = data.get("token")
        if not token:
            raise PcfAuthError("Token mancante nella risposta")

        return self._memorizza_token(token)

    async def _json_or_raise(self, resp: aiohttp.ClientResponse) -> dict:
        """Legge il body JSON (anche in caso di errore: il manuale conferma che
        tutte le risposte, comprese quelle 4xx, hanno body JSON con esito/message)
        e solleva l'eccezione specifica in base allo status HTTP documentato.

        Legge sempre prima il testo grezzo e lo logga (a DEBUG): quando il WAF
        interviene la risposta non è JSON ma HTML, e il testo esatto con il
        Support ID è l'informazione utile per il ticket di assistenza.
        """
        import json as _json

        testo = await resp.text()
        _log_risposta(str(resp.url), resp.status, resp.headers, testo)

        try:
            data = _json.loads(testo)
        except ValueError:
            raise PcfApiError(
                f"Risposta non JSON (HTTP {resp.status}): {testo[:300]}", http_status=resp.status
            )

        if resp.status == 200:
            return data

        message = data.get("message") or f"HTTP {resp.status}"
        if resp.status == 401:
            raise PcfAuthError(message, http_status=401, data=data)
        if resp.status == 409:
            raise PcfConflictError(message, http_status=409, data=data)
        if resp.status == 429:
            raise PcfRateLimitError(message, http_status=429, data=data)
        if resp.status == 404:
            raise PcfNotFoundError(message, http_status=404, data=data)
        if resp.status == 400:
            raise PcfValidationError(message, http_status=400, data=data)
        raise PcfApiError(message, http_status=resp.status, data=data)

    async def async_valida_pod(self, pod: str, df: str) -> str:
        """Verifica che POD e dato fiscale siano validi, restituendo il ticket.

        Fa una vera requestExport su un giorno solo (vedi
        RITARDO_VERIFICA_POD_GIORNI): se la coppia POD/dato
        fiscale non è valida Duereti risponde 400 con "Non ci sono POD/PDR
        validi" (verificato), quindi l'errore emerge subito in fase di
        configurazione invece che al primo ciclo di aggiornamento.

        Il ticket restituito NON va buttato: corrisponde a un job realmente
        accodato lato Duereti. Il chiamante lo salva come pendente, così il
        primo ciclo lo riprende invece di accodarne un altro per lo stesso
        periodo.

        Solleva PcfValidationError se POD o dato fiscale non sono validi,
        PcfAuthError per problemi di credenziali, PcfApiError per il
        resto.
        """
        giorno = date.today() - timedelta(days=RITARDO_VERIFICA_POD_GIORNI)
        return await self.request_export(giorno, giorno, [{"pod": pod, "df": df}])

    async def request_export(
        self,
        data_da: date,
        data_a: date,
        pods: list[dict],
        mode: str = MODE_CURVE,
    ) -> str:
        """Prenota l'estrazione e restituisce il ticket.

        pods: lista di dict {"pod": <codice POD/PDR>, "df": <dato fiscale associato>}
        """
        if len(pods) > MAX_SUPPLY_POINTS_PER_REQUEST:
            raise PcfApiError(
                f"Troppi POD/PDR in una richiesta ({len(pods)}), "
                f"il limite dichiarato dal manuale è {MAX_SUPPLY_POINTS_PER_REQUEST}"
            )
        if (data_a.year - data_da.year) * 12 + (data_a.month - data_da.month) > MAX_DATE_RANGE_MONTHS:
            raise PcfApiError(
                f"Range di date troppo ampio, il limite dichiarato dal manuale è "
                f"{MAX_DATE_RANGE_MONTHS} mesi"
            )

        token = await self._get_token()
        body = {
            # Confermato via test reale: il formato corretto è yyyy-mm-dd
            # (l'esempio JSON del manuale), non dd-mm-yyyy come diceva il
            # testo descrittivo.
            "dataDa": data_da.strftime("%Y-%m-%d"),
            "dataA": data_a.strftime("%Y-%m-%d"),
            "mode": mode,
            "supplyPoints": [{"supplyPoint": p["pod"], "df": p["df"]} for p in pods],
        }
        headers = {"Authorization": token}
        try:
            _log_richiesta(self._url_export, body, headers)
            async with self._session.post(self._url_export, json=body, headers=headers) as resp:
                data = await self._json_or_raise(resp)
        except PcfValidationError as err:
            # Quando esiste già un'elaborazione in corso per gli stessi dati,
            # Duereti risponde 400 e restituisce il ticket di quella esistente
            # in un campo dedicato della risposta (verificato):
            #   {"esito": 1, "message": "vi è già una richiesta ...",
            #    "ticket": "PNbFjolxZVgoAnzA7v8bJw"}
            # Riusarlo evita di accodare un doppione per lo stesso periodo.
            ticket_esistente = (err.data or {}).get("ticket")
            if ticket_esistente:
                _LOGGER.info(
                    "requestExport: elaborazione già in corso per lo stesso periodo, "
                    "riuso il ticket esistente %s",
                    ticket_esistente,
                )
                return ticket_esistente
            raise

        if data.get("esito") != ESITO_OK:
            raise PcfApiError(data.get("message") or "requestExport fallita")

        ticket = data.get("ticket")
        if not ticket:
            raise PcfApiError("Ticket mancante nella risposta di requestExport")
        return ticket

    async def request_result(self, ticket: str) -> bytes:
        """Esegue il polling su requestResult finché il file non è pronto.

        Confermato via test reale: mentre il job è in coda la risposta è
        {"esito": 1, "message": "Il file non è ancora disponibile"} - lo
        trattiamo quindi come "non pronto ancora" e ritentiamo ogni
        RESULT_POLL_INTERVAL_SECONDS (default 10 minuti) fino a
        RESULT_POLL_MAX_ATTEMPTS (default ~6 ore totali). Il token, valido
        10 minuti, viene rinnovato automaticamente ad ogni tentativo.
        """
        token_gia_rinnovato = False

        for attempt in range(1, RESULT_POLL_MAX_ATTEMPTS + 1):
            # Il token dura 10 minuti quanto l'intervallo di polling: lo
            # richiediamo/rinnoviamo ad ogni tentativo (_get_token usa la
            # cache se ancora valido, quindi qui è economico).
            token = await self._get_token()
            headers = {"Authorization": token}
            body = {"ticket": ticket}

            try:
                _log_richiesta(self._url_result, body, headers)
                async with self._session.post(self._url_result, json=body, headers=headers) as resp:
                    data = await self._json_or_raise(resp)
            except PcfAuthError as err:
                # Token rifiutato (verificato: "Token non valido"). Di norma è
                # semplicemente scaduto: forziamo un rinnovo e ritentiamo
                # subito, senza consumare un intero intervallo di attesa.
                if token_gia_rinnovato:
                    # Già rinnovato al giro precedente e ancora rifiutato: non
                    # è una scadenza. Inutile insistere per ore, il polling
                    # terminerebbe con un fuorviante "file non disponibile".
                    raise
                _LOGGER.debug("requestResult: token rifiutato (401), forzo il rinnovo e ritento")
                token_gia_rinnovato = True
                await self._get_token(force_refresh=True)
                continue
            except PcfRateLimitError as err:
                _LOGGER.warning("requestResult: limite richieste raggiunto (429), attendo di più (%s)", err)
                await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS * 2)
                continue
            except PcfNotFoundError as err:
                # Ticket non collegato a nessun dato: errore permanente, non ha
                # senso continuare il polling.
                # Ticket presente ma non collegato a nessun dato: errore
                # permanente, non ha senso continuare il polling. Rilanciamo il
                # tipo originale (non la classe base) così il chiamante può
                # distinguere "ticket davvero invalido" - l'unico caso in cui
                # ha senso scartarlo - da un errore transitorio.
                # Duereti usa lo stesso 404 per due situazioni diverse
                # (entrambe verificate):
                #   "Ticket non trovato"  -> il ticket non esiste
                #   "Nessun dato trovato" -> il ticket è valido, ma per quel
                #                            periodo non ci sono dati
                # In entrambi i casi il polling si ferma, ma il messaggio deve
                # dire la cosa giusta: segnalare "ticket non trovato" quando
                # invece il periodo era vuoto manderebbe a cercare un problema
                # inesistente.
                if "dato" in str(err).lower() or "dati" in str(err).lower():
                    messaggio = (
                        f"Nessun dato disponibile per il periodo richiesto "
                        f"(ticket {ticket}): verifica che le date non siano "
                        f"future o troppo recenti"
                    )
                else:
                    messaggio = f"Ticket {ticket} non trovato lato distributore"
                raise PcfNotFoundError(
                    messaggio,
                    http_status=404,
                    data=getattr(err, "data", None),
                ) from err
            except PcfApiError as err:
                if type(err) is not PcfApiError:
                    # Sottoclasse nota non gestita sopra (es. 400 di
                    # validazione): è un errore applicativo vero, non c'entra
                    # il WAF, non ha senso continuare a ritentare.
                    raise
                # PcfApiError "puro" (nessuna sottoclasse): tipicamente una
                # risposta non-JSON, il sintomo del blocco WAF (vedi
                # _json_or_raise). Non è un segnale sullo stato del ticket -
                # Duereti potrebbe benissimo starlo ancora processando
                # normalmente. Lo trattiamo come "non ancora pronto" e
                # continuiamo il polling con lo STESSO ticket, invece di
                # scartarlo e perdere il lavoro già fatto da Duereti.
                _LOGGER.warning(
                    "requestResult tentativo %s/%s: possibile blocco WAF (%s), continuo il "
                    "polling con lo stesso ticket invece di scartarlo",
                    attempt,
                    RESULT_POLL_MAX_ATTEMPTS,
                    err,
                )
                await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS)
                continue

            if data.get("esito") == ESITO_OK and data.get("File"):
                return self._decode_file_field(data["File"])

            _LOGGER.debug(
                "requestResult tentativo %s/%s: non ancora pronto (%s)",
                attempt,
                RESULT_POLL_MAX_ATTEMPTS,
                data.get("message"),
            )
            await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS)

        raise PcfApiError(
            f"File non disponibile dopo {RESULT_POLL_MAX_ATTEMPTS} tentativi (ticket={ticket})"
        )

    @staticmethod
    def _decode_file_field(file_field: str) -> bytes:
        """Decodifica il campo 'File' della risposta.

        Confermato dal manuale: 'File' è sempre il Base64 dello zip
        (contenente .csv per CURVE o .xlsx per LETTURE).
        """
        try:
            return base64.b64decode(file_field)
        except Exception as err:  # noqa: BLE001
            raise PcfApiError("Impossibile decodificare il campo File come base64") from err


@dataclass
class LetturaRiga:
    """Una riga di lettura periodica (mode=LETTURE), formato confermato via test reale."""

    data_lettura: "datetime"  # noqa: F821
    tipo_lettura: str  # es. "SALDO"
    valore: float
    tipologia_misura: str  # es. "Energia attiva F3"
    matricola_contatore: str
    tipo_misuratore: str
    energia: str  # es. "Energia attiva", "Energia reattiva", "Potenza massima"
    fascia: str  # es. "F1", "F2", "F3"


def parse_letture_zip(zip_bytes: bytes) -> dict[str, list[LetturaRiga]]:
    """Estrae e parsa i file .xlsx dentro lo zip restituito da Duereti per mode=LETTURE.

    Formato CONFERMATO via test reale (non più ipotesi):
    colonne "Data lettura", "Tipo lettura", "Lettura", "Tipologia Misura",
    "Matricola Contatore", "Tipo Misuratore", "Energia", "Fascia". Il foglio
    Excel è nominato come il POD/PDR richiesto, "Lettura" è il valore
    cumulativo (SALDO) del contatore a quella data, non un delta.

    NOTA: questa funzione non è collegata al coordinator dell'integrazione,
    che per scelta usa solo mode=CURVE (dati a intervalli, più adatti alla
    Energy Dashboard). È qui come riferimento/utility nel caso in futuro
    serva anche l'import delle letture periodiche. Richiede 'openpyxl'
    (import lazy per non renderlo una dipendenza hard dell'integrazione).
    """
    import openpyxl
    from datetime import datetime

    risultati: dict[str, list[LetturaRiga]] = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            with zf.open(name) as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)

            for sheet_name in wb.sheetnames:
                # Il foglio è nominato come il POD/PDR
                pod = sheet_name
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    continue
                idx = {col: i for i, col in enumerate(header)}

                for row in rows:
                    try:
                        data_lettura = datetime.strptime(row[idx["Data lettura"]], "%d/%m/%Y")
                        valore = float(str(row[idx["Lettura"]]).replace(",", "."))
                    except (KeyError, ValueError, TypeError, IndexError):
                        _LOGGER.warning("Riga non riconosciuta nel foglio %s: %s", sheet_name, row)
                        continue

                    risultati.setdefault(pod, []).append(
                        LetturaRiga(
                            data_lettura=data_lettura,
                            tipo_lettura=row[idx["Tipo lettura"]],
                            valore=valore,
                            tipologia_misura=row[idx["Tipologia Misura"]],
                            matricola_contatore=row[idx["Matricola Contatore"]],
                            tipo_misuratore=row[idx["Tipo Misuratore"]],
                            energia=row[idx["Energia"]],
                            fascia=row[idx["Fascia"]],
                        )
                    )

    return risultati


def parse_curve_zip(zip_bytes: bytes) -> dict[str, RisultatoLetture]:
    """Estrae e parsa i file .csv dentro lo zip restituito da Duereti per mode=CURVE.

    Formato CONFERMATO via test reale (non più ipotesi), CSV separato da ';':
        POD;DATA;ORA;FL_ORA_LEGALE;ATTIVA_PRELEVATA;ATTIVA_IMMESSA;
        REATTIVA_CAPACITIVA_IMMESSA;REATTIVA_CAPACITIVA_PRELEVATA;
        REATTIVA_INDUTTIVA_IMMESSA;REATTIVA_INDUTTIVA_PRELEVATA;
        PICCO_PRELEVATA;CONSUMO_PICCO_IMMESSA;TIPO_DATO;

    Esempio riga:
        IT001E00000001;20260501;000000;2;0,025;;;;;0,001;5,004;;E;

    Note sul formato:
    - DATA (yyyymmdd) + ORA (hhmmss) = timestamp, a intervalli di 15 minuti
    - i decimali usano la virgola, non il punto
    - ATTIVA_PRELEVATA è energia CONSUMATA nell'intervallo (già un delta per
      i 15 minuti, non un valore cumulativo) - è il campo che importiamo
    - ATTIVA_IMMESSA è energia immessa in rete (produzione, es. fotovoltaico):
      non ancora importata da questa funzione, TODO se in futuro serve
      gestire anche la produzione
    - i campi reattivi/picco non vengono importati
    """
    from datetime import datetime

    risultati: dict[str, RisultatoLetture] = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith((".csv", ".txt")):
                continue
            with zf.open(name) as f:
                raw = f.read().decode("utf-8-sig", errors="replace")

            reader = csv.DictReader(io.StringIO(raw), delimiter=";")

            for row in reader:
                pod = row.get("POD")
                data_raw = row.get("DATA")
                ora_raw = row.get("ORA")
                valore_raw = row.get("ATTIVA_PRELEVATA")

                if not (pod and data_raw and ora_raw):
                    continue

                try:
                    ts = datetime.strptime(f"{data_raw}{ora_raw}", "%Y%m%d%H%M%S")
                except ValueError:
                    _LOGGER.warning("Timestamp non parsabile: DATA=%s ORA=%s", data_raw, ora_raw)
                    continue

                if not valore_raw:
                    # Intervallo senza consumo registrato (es. solo produzione,
                    # o dato mancante): saltiamo, non è un errore.
                    continue

                try:
                    valore = float(valore_raw.replace(",", "."))
                except ValueError:
                    _LOGGER.warning("Valore ATTIVA_PRELEVATA non parsabile: %s", valore_raw)
                    continue

                risultati.setdefault(pod, RisultatoLetture(pod=pod)).punti.append(
                    CurvaPunto(
                        timestamp=ts,
                        valore_kwh=valore,
                        ora_legale=(row.get("FL_ORA_LEGALE") or "").strip() or None,
                    )
                )

    return risultati
